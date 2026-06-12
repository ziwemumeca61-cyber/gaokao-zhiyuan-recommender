"""把 data/incoming/ 里的「省份志愿填报资料包」zip 批量并入 data/real/。

每个省 zip 内嵌一个 zip，内含「XX录取数据22-25/22-25年全国高校在XX的专业录取分数.xlsx」，
列为：年份/院校名称/院校代码/科类/批次/专业/专业代码/所属专业组/专业备注/选科要求/
录取人数/最低分数/最低位次/学校所在/学校性质/是否985/是否211。

合并策略（与现有山东逐校专业数据并存，不覆盖）：
- 院校按名称跨省去重：已存在则复用其 id，否则生成 "R"+hash 的新 id。
- 专业按 (院校, 专业名) 去重：已存在复用 id，否则生成 school_id+"_"+hash。
- 学科门类用现有 majors.csv 的 名称→门类 多数票兜底（缺省留空）。
- 科类映射到 {综合, 物理, 历史}：综合→综合, 物理类→物理, 历史类→历史,
  理科→物理, 文科→历史；艺术/体育/三校生/蒙授等非普通类整行丢弃。
- 录取记录按 (院校,专业,年份,省,科类) 去重，保留首次出现。

用法： python scripts/import_province_packs.py
"""

from __future__ import annotations

import csv
import hashlib
import io
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from gaokao.data_schema import CATEGORY_RIASEC, DEFAULT_RIASEC  # noqa: E402

INCOMING = ROOT / "data" / "incoming"
REAL = ROOT / "data" / "real"

SUBJ_MAP = {"综合": "综合", "物理类": "物理", "历史类": "历史", "理科": "物理", "文科": "历史"}

# 源列名 -> 标准键（含别名）
COL_ALIAS = {
    "年份": "year", "院校名称": "school", "科类": "subj", "专业": "major",
    "选科要求": "subject_req", "录取人数": "plan", "招生人数": "plan",
    "最低分数": "score", "最低分": "score", "最低位次": "rank", "最低分位": "rank",
    "位次": "rank", "学校所在": "loc", "学校性质": "nature", "是否985": "is985",
    "是否211": "is211",
}


def gbk(name: str) -> str:
    try:
        return name.encode("cp437").decode("gbk")
    except Exception:
        return name


def read_existing(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def open_score_xlsx(zip_path: Path) -> io.BytesIO | None:
    """打开省 zip -> 内嵌 zip -> 返回「专业录取分数」xlsx 的字节流。"""
    zf = zipfile.ZipFile(zip_path)
    inners = [n for n in zf.namelist() if n.endswith(".zip")]
    if not inners:
        return None
    izf = zipfile.ZipFile(io.BytesIO(zf.read(inners[0])))
    for n in izf.namelist():
        if n.endswith(".xlsx") and "专业录取分数" in gbk(n):
            return io.BytesIO(izf.read(n))
    return None


def province_of(zip_path: Path) -> str:
    base = gbk(zip_path.name)
    return base.split("、")[1].split("-")[0] if "、" in base else base


def main() -> int:
    schools = read_existing(REAL / "schools.csv")
    majors = read_existing(REAL / "majors.csv")

    sid_by_name = {r["name"]: r["id"] for r in schools}
    used_ids = {r["id"] for r in schools}
    mid_by_key: dict[tuple[str, str], str] = {(r["school_id"], r["name"]): r["id"] for r in majors}
    used_mids = {r["id"] for r in majors}

    # 名称 -> 门类（多数票）
    name_cat_votes: dict[str, Counter] = defaultdict(Counter)
    for r in majors:
        if r["category"]:
            name_cat_votes[r["name"]][r["category"]] += 1
    cat_by_name = {n: c.most_common(1)[0][0] for n, c in name_cat_votes.items()}

    def new_school_id(name: str) -> str:
        sid = "R" + hashlib.md5(name.encode()).hexdigest()[:6]
        while sid in used_ids:
            sid = "R" + hashlib.md5((name + "#").encode()).hexdigest()[:6]
        used_ids.add(sid)
        return sid

    def new_major_id(school_id: str, name: str) -> str:
        mid = f"{school_id}_{hashlib.md5(name.encode()).hexdigest()[:8]}"
        salt = 0
        while mid in used_mids:
            salt += 1
            mid = f"{school_id}_{hashlib.md5((name + str(salt)).encode()).hexdigest()[:8]}"
        used_mids.add(mid)
        return mid

    # 流式写录取记录：先拷贝已有，再追加新省
    adm_fields = ["school_id", "major_id", "year", "province", "subject_type",
                  "min_score", "min_rank", "plan_count"]
    seen: set[tuple] = set()
    tmp_adm = REAL / "admission_scores.tmp.csv"
    adm_f = tmp_adm.open("w", encoding="utf-8-sig", newline="")
    adm_w = csv.DictWriter(adm_f, fieldnames=adm_fields)
    adm_w.writeheader()
    kept_existing = 0
    for r in read_existing(REAL / "admission_scores.csv"):
        key = (r["school_id"], r["major_id"], r["year"], r["province"], r["subject_type"])
        if key in seen:
            continue
        seen.add(key)
        adm_w.writerow(r)
        kept_existing += 1

    new_schools: dict[str, dict] = {}
    new_majors: dict[str, dict] = {}
    per_prov: Counter = Counter()

    zips = sorted(INCOMING.glob("*.zip"))
    for zp in zips:
        prov = province_of(zp)
        bio = open_score_xlsx(zp)
        if bio is None:
            print(f"  {prov}: ✗ 未找到专业录取分数 xlsx，跳过", flush=True)
            continue
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        idx = {COL_ALIAS[h]: i for i, h in enumerate(header) if h in COL_ALIAS}
        need = {"year", "school", "subj", "major", "score", "rank"}
        if not need <= idx.keys():
            print(f"  {prov}: ✗ 缺列 {need - idx.keys()}，跳过", flush=True)
            wb.close()
            continue

        def cell(row, key, default=None):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else default

        n_prov = 0
        for row in rows:
            raw_subj = str(cell(row, "subj") or "").strip()
            subj = SUBJ_MAP.get(raw_subj)
            if subj is None:
                continue
            sname = str(cell(row, "school") or "").strip()
            mname = str(cell(row, "major") or "").strip()
            if not sname or not mname:
                continue
            try:
                year = int(float(cell(row, "year")))
                score = int(float(cell(row, "score")))
                rank = int(float(cell(row, "rank")))
            except (TypeError, ValueError):
                continue
            if rank <= 0:
                continue
            try:
                plan = int(float(cell(row, "plan")))
            except (TypeError, ValueError):
                plan = 0

            # 院校
            sid = sid_by_name.get(sname)
            if sid is None:
                sid = new_school_id(sname)
                sid_by_name[sname] = sid
            if sid not in {s["id"] for s in schools} and sid not in new_schools:
                is985 = str(cell(row, "is985") or "").strip() in ("是", "1", "985")
                is211 = str(cell(row, "is211") or "").strip() in ("是", "1", "211")
                level = "985" if is985 else ("211" if is211 else "普通")
                loc = str(cell(row, "loc") or "").strip()
                nature = str(cell(row, "nature") or "").strip()
                new_schools[sid] = {
                    "id": sid, "name": sname, "province": loc, "city": loc,
                    "level": level, "type": "综合", "tags": nature,
                }

            # 专业
            mkey = (sid, mname)
            mid = mid_by_key.get(mkey)
            if mid is None:
                mid = new_major_id(sid, mname)
                mid_by_key[mkey] = mid
                category = cat_by_name.get(mname, "")
                new_majors[mid] = {
                    "id": mid, "name": mname, "category": category, "school_id": sid,
                    "riasec_code": CATEGORY_RIASEC.get(category, DEFAULT_RIASEC),
                    "heat": 50.0, "employment_rate": 0.85,
                    "subject_req": str(cell(row, "subject_req") or "").strip(),
                    "intro": "", "core_courses": "", "career_paths": "",
                    "industry_outlook": "", "suits": "",
                }

            akey = (sid, mid, str(year), prov, subj)
            if akey in seen:
                continue
            seen.add(akey)
            adm_w.writerow({
                "school_id": sid, "major_id": mid, "year": year, "province": prov,
                "subject_type": subj, "min_score": score, "min_rank": rank,
                "plan_count": max(plan, 0),
            })
            n_prov += 1
        wb.close()
        per_prov[prov] = n_prov
        print(f"  {prov}: +{n_prov} 条录取记录", flush=True)

    adm_f.close()

    # 写院校 / 专业（已有 + 新增）
    sch_fields = ["id", "name", "province", "city", "level", "type", "tags"]
    maj_fields = ["id", "name", "category", "school_id", "riasec_code", "heat",
                  "employment_rate", "subject_req", "intro", "core_courses",
                  "career_paths", "industry_outlook", "suits"]
    with (REAL / "schools.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=sch_fields)
        w.writeheader()
        for r in schools:
            w.writerow({k: r.get(k, "") for k in sch_fields})
        for r in new_schools.values():
            w.writerow(r)
    with (REAL / "majors.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=maj_fields)
        w.writeheader()
        for r in majors:
            w.writerow({k: r.get(k, "") for k in maj_fields})
        for r in new_majors.values():
            w.writerow(r)
    tmp_adm.replace(REAL / "admission_scores.csv")

    print("\n=== 完成 ===")
    print(f"新增院校 {len(new_schools)}，新增专业 {len(new_majors)}")
    print(f"录取记录：已有 {kept_existing} + 新增 {sum(per_prov.values())}"
          f" = {kept_existing + sum(per_prov.values())}")
    print("各省新增：", dict(per_prov))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
