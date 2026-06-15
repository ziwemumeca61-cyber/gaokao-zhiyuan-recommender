"""用真实「招生计划」修正 data/real 的 plan_count（扩招/缩招信号更准）。

现状：plan_count 此前由「专业录取分数」表的 *录取人数* 填充，但 recommender/history.py
把它当 *招生计划数* 用来算 plan_ratio（扩招→分数线走低）。本脚本读取各省资料包里的
「招生计划」xlsx（列含 年份/院校名称/科类/专业名称/招生人数/学费(元)/学制(年)），
按 (院校,专业,年份,省,科类) 汇总真实计划数，覆盖匹配到的 admission_scores.plan_count。

匹配规则与 import_province_packs 一致（按名称对到已有 school_id / major_id）；同一
(校,专业,年,省,科类) 的多个专业组/批次计划数求和。未匹配的记录保持原值（优雅降级）。

用法： python scripts/import_plans.py [incoming_dir]
"""

from __future__ import annotations

import csv
import gzip
import io
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REAL = ROOT / "data" / "real"
ADM_GZ = REAL / "admission_scores.csv.gz"

SUBJ_MAP = {"综合": "综合", "物理类": "物理", "历史类": "历史", "理科": "物理", "文科": "历史"}


def gbk(name: str) -> str:
    for enc in ("gbk", "utf-8"):
        try:
            return name.encode("cp437").decode(enc)
        except Exception:  # noqa: BLE001
            pass
    return name


def province_of(zip_path: Path) -> str:
    base = gbk(zip_path.name)
    return base.split("、")[1].split("-")[0] if "、" in base else base


def open_plan_xlsx(zip_path: Path) -> io.BytesIO | None:
    zf = zipfile.ZipFile(zip_path)
    inners = [n for n in zf.namelist() if n.endswith(".zip")]
    container = zipfile.ZipFile(io.BytesIO(zf.read(inners[0]))) if inners else zf
    for n in container.namelist():
        g = gbk(n)
        if n.endswith(".xlsx") and "招生计划" in g and "22-25" in g:
            return io.BytesIO(container.read(n))
    return None


def read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    incoming = Path(args[0]) if args else (ROOT / "data" / "incoming")

    schools = read_csv(REAL / "schools.csv")
    majors = read_csv(REAL / "majors.csv")
    sid_by_name = {r["name"]: r["id"] for r in schools}
    mid_by_key = {(r["school_id"], r["name"]): r["id"] for r in majors}

    # (sid, mid, year, prov, subj) -> 计划数合计
    plan_sum: dict[tuple, int] = defaultdict(int)
    for zp in sorted(incoming.glob("*.zip")):
        prov = province_of(zp)
        bio = open_plan_xlsx(zp)
        if bio is None:
            print(f"  {prov}: ✗ 无招生计划表，跳过", flush=True)
            continue
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        h = [str(x).strip() if x is not None else "" for x in next(rows)]
        col = {name: h.index(name) for name in
               ("年份", "院校名称", "科类", "专业名称", "招生人数") if name in h}
        if len(col) < 5:
            print(f"  {prov}: ✗ 缺列，跳过", flush=True)
            wb.close()
            continue
        n = 0
        for r in rows:
            subj = SUBJ_MAP.get(str(r[col["科类"]]).strip() if r[col["科类"]] else "")
            if subj is None:
                continue
            sid = sid_by_name.get(str(r[col["院校名称"]] or "").strip())
            if sid is None:
                continue
            mid = mid_by_key.get((sid, str(r[col["专业名称"]] or "").strip()))
            if mid is None:
                continue
            try:
                year = int(float(r[col["年份"]]))
                seats = int(float(r[col["招生人数"]]))
            except (TypeError, ValueError):
                continue
            if seats <= 0:
                continue
            plan_sum[(sid, mid, str(year), prov, subj)] += seats
            n += 1
        wb.close()
        print(f"  {prov}: 匹配 {n} 条计划行", flush=True)

    # 重写 admission_scores：覆盖匹配到的 plan_count
    rows = []
    with gzip.open(ADM_GZ, "rt", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames
        updated = 0
        for rec in reader:
            key = (rec["school_id"], rec["major_id"], rec["year"],
                   rec["province"], rec["subject_type"])
            if key in plan_sum:
                rec["plan_count"] = plan_sum[key]
                updated += 1
            rows.append(rec)

    with gzip.open(ADM_GZ, "wt", encoding="utf-8-sig", newline="", compresslevel=9) as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    print(f"\n=== 完成 ===\n汇总真实计划键 {len(plan_sum):,}；"
          f"覆盖更新 admission_scores 记录 {updated:,} / {len(rows):,}。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
