"""从 data/incoming/ 的省份资料包提取「一分一段表」种子到 data/segments/。

每个省 zip（或其内嵌 zip）里有 `.../一分一段/XX{年}年的一分一段表.xlsx`，列为：
年份/科类/批次/控制线(分)/分数(分)/本段人数(人)/累计人数(人)/排名区间/历史同位次考生得分。
取**最新年份**那份，按科类拆出 (分数, 累计人数=位次) 写成 `{省}_{科类}_{年}.csv`
（列 `score,rank`，与 gaokao.rank_score 读取格式一致）。

约定：
- 科类映射到 {综合, 物理, 历史}：综合→综合, 物理类/理科→物理, 历史类/文科→历史；其余丢弃。
- 分数段形如 "692-750"（最高段）取下界 692；单值直接用。
- 默认**不覆盖**已存在的真实种子（data/segments 里同名 (省,科类)）——用 --force 才覆盖。
- 同步在 data/segments/sources.json 写入出处。

用法：
    python scripts/extract_segments.py [incoming_dir] [--force]
"""

from __future__ import annotations

import csv
import io
import json
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SEGMENTS = ROOT / "data" / "segments"

SUBJ_MAP = {"综合": "综合", "物理类": "物理", "历史类": "历史", "理科": "物理", "文科": "历史"}

# 招生考试机构全称（自治区/直辖市不能简单加“省”）
AUTHORITY = {
    "新疆": "新疆维吾尔自治区教育考试院", "广西": "广西壮族自治区招生考试院",
    "宁夏": "宁夏回族自治区教育考试院", "内蒙古": "内蒙古自治区教育招生考试中心",
    "北京": "北京教育考试院", "天津": "天津市教育招生考试院", "上海": "上海市教育考试院",
    "重庆": "重庆市教育考试院",
}


def authority_of(prov: str) -> str:
    return AUTHORITY.get(prov, f"{prov}省教育考试院")


def gbk(name: str) -> str:
    for enc in ("gbk", "utf-8"):
        try:
            return name.encode("cp437").decode(enc)
        except Exception:  # noqa: BLE001, PERF203
            pass
    return name


def province_of(zip_path: Path) -> str:
    base = gbk(zip_path.name)
    return base.split("、")[1].split("-")[0] if "、" in base else base


def latest_segment_xlsx(zip_path: Path) -> tuple[str, io.BytesIO] | None:
    """返回 (年份, 最新「一分一段」xlsx 字节流)。兼容内嵌 zip / 外层直放两种结构。"""
    zf = zipfile.ZipFile(zip_path)
    inners = [n for n in zf.namelist() if n.endswith(".zip")]
    container = zipfile.ZipFile(io.BytesIO(zf.read(inners[0]))) if inners else zf
    best: tuple[str, str] | None = None
    for n in container.namelist():
        g = gbk(n)
        if n.endswith(".xlsx") and "/一分一段/" in g and "年的一分一段" in g:
            digits = "".join(c for c in g.split("/")[-1] if c.isdigit())[:4]
            if best is None or digits > best[0]:
                best = (digits, n)
    if best is None:
        return None
    return best[0], io.BytesIO(container.read(best[1]))


def parse_score(cell: object) -> int | None:
    """'692-750' -> 692（下界）；'697' -> 697；其它 -> None。"""
    s = str(cell).strip()
    if not s:
        return None
    nums = [int(x) for x in s.replace("－", "-").split("-") if x.strip().isdigit()]
    return min(nums) if nums else None


def extract_one(zip_path: Path) -> dict[str, list[tuple[int, int]]]:
    """返回 {映射科类: [(score, rank), ...]}（取最新年份）。"""
    prov = province_of(zip_path)
    found = latest_segment_xlsx(zip_path)
    if found is None:
        print(f"  {prov}: ✗ 无一分一段，跳过", flush=True)
        return {}
    year, bio = found
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    try:
        si = header.index("分数(分)")
        ri = header.index("累计人数(人)")
        ki = header.index("科类")
    except ValueError:
        print(f"  {prov}: ✗ 缺关键列，跳过", flush=True)
        wb.close()
        return {}
    out: dict[str, list[tuple[int, int]]] = {}
    for row in rows:
        subj = SUBJ_MAP.get(str(row[ki]).strip() if ki < len(row) and row[ki] else "")
        if subj is None:
            continue
        score = parse_score(row[si]) if si < len(row) else None
        try:
            rank = int(float(row[ri]))
        except (TypeError, ValueError):
            rank = 0
        if score is None or rank <= 0:
            continue
        out.setdefault(subj, []).append((score, rank))
    wb.close()
    out["_year"] = year  # type: ignore[assignment]
    return out


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    force = "--force" in sys.argv[1:]
    incoming = Path(args[0]) if args else (ROOT / "data" / "incoming")

    SEGMENTS.mkdir(parents=True, exist_ok=True)
    sources_path = SEGMENTS / "sources.json"
    sources = json.loads(sources_path.read_text(encoding="utf-8")) if sources_path.exists() else {}

    written, skipped = 0, 0
    for zp in sorted(incoming.glob("*.zip")):
        data = extract_one(zp)
        year = data.pop("_year", "")  # type: ignore[arg-type]
        prov = province_of(zp)
        for subj, pairs in data.items():
            key = f"{prov}_{subj}_{year}"
            out_csv = SEGMENTS / f"{key}.csv"
            # 不覆盖任何已存在的同 (省,科类) 真实种子
            existing = list(SEGMENTS.glob(f"{prov}_{subj}_*.csv"))
            if existing and not force:
                print(f"  {prov}/{subj}: 已有种子 {existing[0].name}，跳过", flush=True)
                skipped += 1
                continue
            # 按分数升序去重（同分取最大累计=最差位次），交给加载器做单调清洗
            best_by_score: dict[int, int] = {}
            for s, r in pairs:
                best_by_score[s] = max(best_by_score.get(s, 0), r)
            ordered = sorted(best_by_score.items())
            with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(["score", "rank"])
                w.writerows(ordered)
            sources[key] = {
                "province": prov, "subject_type": subj, "year": int(year) if year.isdigit() else year,
                "source": authority_of(prov),
                "title": f"{year}年普通高考一分一段统计表（{subj}类）",
                "url": "",
                "note": "官方成绩分段统计（公开统计事实），由志愿填报资料包整理；"
                        "分数段之间按对数位次线性插值，超出区间按趋势外推并标注为估算。",
            }
            print(f"  {prov}/{subj}: +{len(ordered)} 段 -> {out_csv.name}", flush=True)
            written += 1

    sources_path.write_text(
        json.dumps(sources, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n=== 完成 ===\n写入 {written} 份，跳过 {skipped} 份；sources.json 已更新。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
